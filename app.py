import streamlit as st
import pandas as pd
import os
import io
import openpyxl
from openpyxl.styles import Border, Side, Alignment

# 1. Configuración de pantalla en modo ancho
st.set_page_config(page_title="Control O&M Semanal", layout="wide")

# ENLACE DIRECTO DE TU HOJA DE GOOGLE (Para lectura pública)
URL_HOJA_DIRECTA = "https://docs.google.com/spreadsheets/d/1veXDuDIPG7KhUM3drp034Y7NfdY-3I0NyAqhRtRiFt4/edit"
CSV_URL = "https://docs.google.com/spreadsheets/d/1veXDuDIPG7KhUM3drp034Y7NfdY-3I0NyAqhRtRiFt4/export?format=csv"

# INYECCIÓN CSS AVANZADA: ELIMINAR BARRA DE DESARROLLO COMPLETAMENTE
st.markdown("""
    <style>
        div[data-testid="stStatusWidget"] { display: none !important; }
        [data-testid="stDecoration"] { display: none !important; }
        [data-testid="stActionButton"] { display: none !important; }
        header { visibility: hidden !important; display: none !important; height: 0px !important; }
        #MainMenu { visibility: hidden !important; }
        footer { visibility: hidden !important; }
        .main .block-container { padding-top: 1rem !important; }
        [data-testid="stSidebarCollapseButton"] { display: none !important; visibility: hidden !important; }
    </style>
""", unsafe_allow_html=True)

# Mantener los datos persistentes en la sesión de la app
if "df_datos" not in st.session_state:
    st.session_state.df_datos = pd.DataFrame()
if "mostrar_globos" not in st.session_state:
    st.session_state.mostrar_globos = False

if st.session_state.mostrar_globos:
    st.balloons()
    st.session_state.mostrar_globos = False

# ACCESO DIRECTO DE ADMINISTRACIÓN
with st.expander("🛠️ Panel de Acceso - Administrador (Dar clic aquí para cargar Excel)"):
    clave_directa = st.text_input("Introduce la clave de acceso de O&M", type="password", key="clave_main")
    
    if clave_directa == "admin123":
        st.success("👨‍💻 Modo Administrator Activo")
        st.write("---")
        st.subheader("Cargar Nueva Semana Activa")
        nuevo_excel_main = st.file_uploader("Subir archivo Excel (.xlsx)", type=["xlsx"], key="uploader_main")
        
        if nuevo_excel_main is not None:
            if st.button("🔄 Inicializar Nueva Semana", use_container_width=True):
                try:
                    # Guardar la plantilla en el servidor local
                    with open("plantilla_original.xlsx", "wb") as f:
                        f.write(nuevo_excel_main.getbuffer())

                    df_t = pd.read_excel(nuevo_excel_main, header=None, dtype=str).fillna("")
                    fila_h = 0
                    for i, r in df_t.iterrows():
                        vals = [str(v).strip().upper() for v in r.values]
                        if any("SITIO" in x or "SEMANA" in x for x in vals):
                            fila_h = i
                            break
                    
                    df_c = pd.read_excel(nuevo_excel_main, skiprows=fila_h, dtype=str).fillna("")
                    df_c.columns = df_c.columns.str.strip()
                    columnas_reales = list(df_c.columns)
                    
                    if len(columnas_reales) >= 8: columnas_reales[7] = "Se realizó (Si/No)"
                    if len(columnas_reales) >= 9: columnas_reales[8] = "Si no se realizó detallar el por qué."
                    df_c.columns = columnas_reales
                    
                    df_c = df_c[df_c["SITIO"] != ""]
                    df_c = df_c[~df_c["SITIO"].str.contains("Total", na=False, case=False)]
                    df_c.loc[df_c["Se realizó (Si/No)"] == "", "Se realizó (Si/No)"] = "Pendiente"
                    
                    # Forzar el almacenamiento local en memoria de la App sin depender de permisos de escritura de Google
                    st.session_state.df_datos = df_c
                    df_c.to_csv("copia_seguridad.csv", index=False)
                    
                    st.success("¡Estructura oficial guardada en memoria local correctamente!")
                    st.rerun() 
                except Exception as e:
                    st.error(f"Error al procesar: {e}")
    elif clave_directa != "":
        st.error("❌ Contraseña incorrecta.")

st.write("---")
st.title("📊 RIR/RDA Status Tracker - O&M")

# CARGAR DATOS INTELIGENTES
if st.session_state.df_datos.empty:
    if os.path.exists("copia_seguridad.csv"):
        st.session_state.df_datos = pd.read_csv("copia_seguridad.csv", dtype=str).fillna("")
    else:
        try:
            # Intenta leer de forma pública tu hoja de Google Sheets
            st.session_state.df_datos = pd.read_csv(CSV_URL, dtype=str).fillna("")
        except Exception:
            st.session_state.df_datos = pd.DataFrame()

# INTERFAZ PRINCIPAL DE TRABAJO
if not st.session_state.df_datos.empty and "SITIO" in st.session_state.df_datos.columns:
    df_actual = st.session_state.df_datos
    num_sem = df_actual['SEMANA'].iloc[0] if 'SEMANA' in df_actual.columns else ""
    st.subheader(f"📋 Sitios Activos - Semana {num_sem}")
    
    df_editado = st.data_editor(
        df_actual,
        column_config={
            "Se realizó (Si/No)": st.column_config.SelectboxColumn("Se realizó (Si/No)", options=["Pendiente", "Si", "No", "Cancelado"], required=True),
            "Si no se realizó detallar el por qué.": st.column_config.TextColumn("Si no se realizó detallar el por qué.", width="large")
        },
        disabled=[c for c in df_actual.columns if c not in ["Se realizó (Si/No)", "Si no se realizó detallar el por qué."]],
        hide_index=True, use_container_width=True
    )

    if st.button("💾 Guardar Mis Cambios del Día", use_container_width=True):
        st.session_state.df_datos = df_editado
        df_editado.to_csv("copia_seguridad.csv", index=False)
        st.session_state.mostrar_globos = True
        st.success("¡Tus cambios del día se guardaron con éxito en el sistema!")
        st.rerun()

    st.write("---")
    total = len(df_editado)
    hechos = len(df_editado[df_editado["Se realizó (Si/No)"] == "Si"])
    
    col1, col2 = st.columns([2, 1])
    with col1:
        st.write(f"**Avance del Equipo:** {hechos} de {total} completados.")
        st.progress(hechos / total if total > 0 else 0.0)
    with col2:
        if os.path.exists("plantilla_original.xlsx"):
            try:
                wb = openpyxl.load_workbook("plantilla_original.xlsx")
                ws = wb.active
                idx_s, fila_t = None, None
                for r in range(1, 15):
                    vals = [str(ws.cell(row=r, column=c).value).strip().upper() for c in range(1, ws.max_column + 1)]
                    if any("SITIO" in v or "SEMANA" in v for v in vals):
                        fila_t = r
                        for c in range(1, ws.max_column + 1):
                            if "SITIO" in str(ws.cell(row=r, column=c).value).strip().upper(): idx_s = c
                        break
                
                if idx_s:
                    map_r = dict(zip(df_editado["SITIO"], df_editado["Se realizó (Si/No)"]))
                    map_j = dict(zip(df_editado["SITIO"], df_editado["Si no se realizó detallar el por qué."]))
                    borde = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
                    
                    for f in range(fila_t + 1, ws.max_row + 1):
                        val_s = ws.cell(row=f, column=idx_s).value
                        if val_s and str(val_s).strip() in map_r:
                            cod_s = str(val_s).strip()
                            c_r = ws.cell(row=f, column=8, value=map_r[cod_s])
                            c_j = ws.cell(row=f, column=9, value=map_j[cod_s])
                            c_r.border, c_j.border = borde, borde
                            c_r.alignment = Alignment(horizontal="center", vertical="center")
                            c_j.alignment = Alignment(horizontal="left", vertical="center")
                    
                    buf = io.BytesIO()
                    wb.save(buf)
                    
                    st.download_button(
                        label=f"📥 Descargar Reporte Semana {num_sem} (.xlsx)", 
                        data=buf.getvalue(), 
                        file_name=f"Reporte_O&M_Semana_{num_sem}.xlsx", 
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                        use_container_width=True
                    )
            except Exception as ex:
                st.error(f"Error al preparar la descarga: {ex}")
else:
    st.warning("⏳ Esperando que el administrador cargue los datos iniciales de la semana...")
